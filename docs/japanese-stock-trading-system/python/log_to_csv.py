#!/usr/bin/env python3
"""Dump Excel log sheets to CSV / SQLite after market close.

Run from VBA at 15:30:
    pythonw.exe log_to_csv.py --date 2026-04-09

Sheets dumped:
    - 売買ログ   -> trade_log.csv / trade_log table
    - 約定履歴   -> fills.csv     / fills table
    - エラーログ -> errors.csv    / errors table
"""
from __future__ import annotations

import argparse
import csv
import sqlite3
from datetime import date
from pathlib import Path

import openpyxl  # pip install openpyxl
import yaml      # pip install pyyaml

ROOT = Path(__file__).resolve().parent
CONFIG_PATH = ROOT / "config.yaml"

SHEETS = {
    "売買ログ":   ("trade_log", "trade_log.csv"),
    "約定履歴":   ("fills",     "fills.csv"),
    "エラーログ": ("errors",    "errors.csv"),
}


def load_config() -> dict:
    return yaml.safe_load(CONFIG_PATH.read_text(encoding="utf-8"))


def dump_to_csv(wb_path: Path, sheet_name: str, out_path: Path) -> None:
    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(row)


def dump_to_sqlite(wb_path: Path, sheet_name: str, table: str, db: Path) -> None:
    wb = openpyxl.load_workbook(wb_path, data_only=True, read_only=True)
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return
    header = [str(c) if c is not None else f"col{i}"
              for i, c in enumerate(rows[0])]
    db.parent.mkdir(parents=True, exist_ok=True)
    with sqlite3.connect(db) as con:
        cols_sql = ", ".join(f'"{h}" TEXT' for h in header)
        con.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({cols_sql});')
        placeholders = ",".join(["?"] * len(header))
        q = f'INSERT INTO "{table}" VALUES ({placeholders});'
        con.executemany(q, [
            tuple(str(v) if v is not None else None for v in r)
            for r in rows[1:]
        ])
        con.commit()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", default=date.today().isoformat())
    args = parser.parse_args()

    cfg = load_config()
    wb_path = Path(cfg["excel"]["workbook"])
    log_root = Path(cfg["logs"]["root_dir"])
    fmt = cfg["logs"].get("format", "both")
    db = Path(cfg["logs"]["sqlite_path"]) if "sqlite" in (fmt, "both") else None

    d = args.date
    target = log_root / d[0:4] / d[5:7] / d[8:10]

    for sheet, (table, filename) in SHEETS.items():
        try:
            dump_to_csv(wb_path, sheet, target / filename)
            if db is not None and fmt in ("sqlite", "both"):
                dump_to_sqlite(wb_path, sheet, table, db)
        except Exception as e:
            (ROOT / "python_errors.log").open("a", encoding="utf-8").write(
                f"log_to_csv\t{sheet}\t{e}\n")
            return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
